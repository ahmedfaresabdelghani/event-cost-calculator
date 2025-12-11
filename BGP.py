#!/usr/bin/env python3
# bfd_to_excel.py
# Requirements: pip install netmiko openpyxl

import re
import os
import time
from datetime import datetime
from collections import defaultdict, OrderedDict

from getpass import getpass
from tqdm import tqdm

from netmiko import ConnectHandler, NetMikoTimeoutException, NetMikoAuthenticationException
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ------------------------- CONFIG -------------------------
NODES = [
    ("TNT-02ASR02_CI-02", "10.204.64.6"),
    ("TNT-02ASR01_CI-01", "10.204.64.9"),
    ("MAN-09ASR01_CI-01", "10.205.64.9"),
    ("MAN-09ASR02_CI-02", "10.205.64.6"),
    ("BS-08ASR01_CI-01", "172.23.172.4"),
    ("BS-08ASR02_CI-02", "172.23.172.7"),
    ("CA5-07ASR01_CI-01", "172.21.41.23"),
    ("CA5-07ASR02_CI-02", "172.21.41.26"),
    ("CA5-07ASR09_CI-09", "10.21.75.100"),
    ("CA5-07ASR10_CI-10", "10.21.75.103"),
    ("MKT-03ASR01_CI-01", "10.19.3.6"),
    ("MKT-03ASR02_CI-02", "10.19.3.9"),
    ("ALX-05ASR01_CI-01", "172.27.202.35"),
    ("ALX-05ASR02_CI-02", "172.27.202.38"),
    ("CA4-06ASR01_CI-01", "172.18.41.3"),
    ("CA4-06ASR02_CI-02", "172.18.41.11"),
    ("CA4-06ASR09_CI-09", "10.18.58.4"),
    ("CA4-06ASR10_CI-10", "10.18.58.7"),
    ("RMD-04ASR01_CI-01", "172.28.41.38"),
    ("RMD-04ASR02_CI-02", "172.28.41.41"),
    ("RMD-04ASR09_CI-09", "10.28.53.196"),
    ("RMD-04ASR10_CI-10", "10.28.53.199"),
    ("HQ-01ASR01_CI-01", "172.30.41.26"),
    ("HQ-01ASR02_CI-02", "172.30.41.29"),
    ("HQ-01ASR09_CI-09", "10.30.78.4"),
    ("HQ-01ASR10_CI-10", "10.30.78.7"),
]

# ==================== LOGIN ====================

USERNAME = input("Enter username: ")
PASSWORD = getpass("Enter password: ")

DEVICE_TYPE = "cisco_xr"

# Commands
LOG_CMD = "show logging start today | i bfd | i BV"
INT_DES_CMD_TEMPLATE = "show int {iface} des"

OUT_DIR = "outputs"
EXCEL_FILE = os.path.join(OUT_DIR, "BFD_Status_Report.xlsx")
TXT_FILE = os.path.join(OUT_DIR, "combined_report.txt")
LOG_FILE = os.path.join(OUT_DIR, "run_log.txt")

os.makedirs(OUT_DIR, exist_ok=True)

# ------------------------- Regex -------------------------
# Capture SESSION_STATE_UP / SESSION_STATE_DOWN (ignore DAMPENING)
RE_STATE = re.compile(r"SESSION_STATE_(UP|DOWN)", re.IGNORECASE)
RE_DAMP = re.compile(r"SESSION_DAMPENING", re.IGNORECASE)
RE_NEIGH = re.compile(r"neighbor\s+(\d+\.\d+\.\d+\.\d+)", re.IGNORECASE)
# Interface may be "BVI123" or "BV123" or "BV123" variants
RE_INTF = re.compile(r"interface\s+(BVI?\d+|BV\d+|BV\d+)", re.IGNORECASE)
# time extraction
RE_TIME = re.compile(r"\b[A-Za-z]{3}\s+\d{1,2}\s+(\d{2}:\d{2}:\d{2}(?:\.\d+)?)\b")

# ------------------------- Company keywords -------------------------
COMPANY_KEYWORDS = {
    "Etisalat": ["etisalat", "eti", "tele", "telemisr", "ettislat", "etislat"],
    "Orange": ["orange", "org"],
    "WE": ["we", "te", "te-fixed", "te-fixed", "te-"],  # WE/TE variants
}
# fallback company name
OTHER_COMPANY_NAME = "OTHER"

# ------------------------- Excel style setup -------------------------
# header fills for companies
FILL_ETISALAT = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
FILL_ORANGE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")    # light orange
FILL_WE = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")        # light purple
FILL_OTHER = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")     # grey
FILL_DOWN = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # red-ish for DOWN
FILL_UP = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")        # green for UP

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BOLD = Font(bold=True)

# ------------------------- Utility functions -------------------------
def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{ts} - {msg}\n")

def classify_company(description: str) -> str:
    if not description:
        return OTHER_COMPANY_NAME
    d = description.lower()
    for comp, kws in COMPANY_KEYWORDS.items():
        for kw in kws:
            if kw in d:
                return comp
    return OTHER_COMPANY_NAME

def parse_log_for_bv_entries(log_text: str):
    """
    Parse the log output (already filtered by | i BV) and return per-interface info:
    returns ordered dict: iface -> dict { 'last_time':..., 'last_state': 'UP'/'DOWN', 'peers': [ips] (unique ordered) }
    We ignore lines with DAMPENING (SESSION_DAMPENING).
    """
    iface_map = OrderedDict()
    for line in log_text.splitlines():
        if not line.strip():
            continue
        if RE_DAMP.search(line):
            continue  # ignore dampening lines
        m_state = RE_STATE.search(line)
        if not m_state:
            continue
        m_intf = RE_INTF.search(line)
        m_neigh = RE_NEIGH.search(line)
        if not m_intf:
            continue
        iface = m_intf.group(1)
        state = m_state.group(1).upper()
        neigh = m_neigh.group(1) if m_neigh else ""
        t_m = RE_TIME.search(line)
        t = t_m.group(1) if t_m else ""
        if iface not in iface_map:
            iface_map[iface] = {"last_time": t, "last_state": state, "peers": []}
        else:
            # update last_time and last_state (we take the latest occurrence in file order,
            # assuming file is chronological; we still update to reflect latest line.)
            iface_map[iface]["last_time"] = t
            iface_map[iface]["last_state"] = state
        if neigh and neigh not in iface_map[iface]["peers"]:
            iface_map[iface]["peers"].append(neigh)
    return iface_map

def parse_interface_description(output_text: str, iface: str):
    """
    Parse 'show int <iface> des' output and get:
      - description string (full)
      - interface status: "Up" or "Down" according to rules:
          if both Status and Protocol columns show 'up' -> Up, else Down.
    The command output format is expected similar to:
    Interface          Status      Protocol    Description
    --------------------------------------------------------------------------------
    BV527              up          up          Etisalat-MKT/RMS
    """
    lines = [l for l in output_text.splitlines() if l.strip()]
    if not lines:
        return ("NO_DESC_FOUND", "UNKNOWN")
    # find a line that contains numeric part of iface
    num = re.sub(r"\D", "", iface)
    for line in lines:
        if num and num in line:
            parts = re.split(r"\s{2,}", line.strip())
            # typical: [iface] [status] [protocol] [description]
            if len(parts) >= 4:
                iface_col = parts[0].strip()
                status_col = parts[1].strip().lower()
                proto_col = parts[2].strip().lower()
                desc_col = parts[3].strip()
                # determine final status: Up only if status_col == 'up' and proto_col == 'up'
                final_status = "Up" if status_col == "up" and proto_col == "up" else "Down"
                return (desc_col if desc_col else "NO_DESC_FOUND", final_status)
            elif len(parts) >= 2:
                # fallback: use last column as desc
                desc_col = parts[-1].strip()
                return (desc_col if desc_col else "NO_DESC_FOUND", "UNKNOWN")
    # fallback: try to get last non-header line
    for line in reversed(lines):
        if re.search(r"-{3,}", line):
            continue
        parts = re.split(r"\s{2,}", line.strip())
        if len(parts) >= 2:
            desc_col = parts[-1].strip()
            # we may not have explicit status/proto; return unknown status
            return (desc_col if desc_col else "NO_DESC_FOUND", "UNKNOWN")
    return ("NO_DESC_FOUND", "UNKNOWN")

# ------------------------- Main processing -------------------------
def process_all_nodes():
    # Prepare aggregate structure for Excel: list of rows per node (we will expand rows)
    report_per_node = OrderedDict()  # node -> dict(company -> list of dicts {iface, peer, status, desc, time}))
    txt_report_blocks = []

    for node_name, node_ip in tqdm(NODES, desc="Processing nodes", unit="node"):
        log(f"Start node {node_name} {node_ip}")
        node_entries = defaultdict(list)  # company -> list of entries
        try:
            device = {
                "device_type": DEVICE_TYPE,
                "host": node_ip,
                "username": USERNAME,
                "password": PASSWORD,
                "port": 22,
                "banner_timeout": 60,
            }
            conn = ConnectHandler(**device)
            # 1) get filtered logs
            logs = conn.send_command(LOG_CMD, expect_string=r"#|>", delay_factor=2, max_loops=600, read_timeout=120)
            # parse BV/BVI lines and last states
            iface_map = parse_log_for_bv_entries(logs)
            if not iface_map:
                # no BVI/BV events
                txt_report_blocks.append(f"{'-'*41}{node_name}{'-'*41}\nNo BGP Flapped / Down\n")
                report_per_node[node_name] = node_entries
                conn.disconnect()
                
                continue

            # cache descriptions per iface
            desc_cache = {}
            status_from_desc = {}  # iface -> Up/Down/UNKNOWN (from show int ... des)
            for iface in iface_map.keys():
                cmd = INT_DES_CMD_TEMPLATE.format(iface=iface)
                try:
                    out = conn.send_command(cmd, expect_string=r"#|>", delay_factor=1, max_loops=200, read_timeout=60)
                except Exception as e:
                    out = ""
                desc, int_status = parse_interface_description(out, iface)
                desc_cache[iface] = desc
                status_from_desc[iface] = int_status

            # Now group by inferred company using desc_cache
            for iface, info in iface_map.items():
                desc = desc_cache.get(iface, "NO_DESC_FOUND")
                company = classify_company(desc)
                entry = {
                    "iface": iface,
                    "peers": info.get("peers", []),
                    "log_state": info.get("last_state", "UNKNOWN"),   # UP/DOWN from logs last line
                    "time": info.get("last_time", ""),
                    "desc": desc,
                    "int_status": status_from_desc.get(iface, "UNKNOWN"),  # Up/Down/UNKNOWN from show int ... des
                }
                node_entries[company].append(entry)

            # Build text blocks similar to earlier format (one block per company per node)
            for comp, entries in node_entries.items():
                # build combined lists per company
                ifaces_str = " , ".join(e["iface"] for e in entries)
                peers_all = []
                for e in entries:
                    for p in e["peers"]:
                        if p not in peers_all:
                            peers_all.append(p)
                peers_str = " , ".join(peers_all)
                descs = " , ".join(e["desc"] for e in entries)
                # pick alarm time as earliest (first) found time in entries (they are in log order)
                alarm_time = entries[0]["time"] if entries and entries[0]["time"] else ""
                # classification logic: if any last_state == DOWN and last occurrence is UP then FLAPPED; we use last_state logic:
                last_states = [e["log_state"] for e in entries]
                last_state = last_states[-1] if last_states else "UNKNOWN"
                any_down = any(s == "DOWN" for s in last_states)
                if last_state == "DOWN":
                    classification = "BGP Down"
                elif last_state == "UP" and any_down:
                    classification = "BGP Flapped"
                else:
                    classification = "BGP Flapped"

                block_lines = [
                    f"{'-'*41}{node_name}{'-'*41}",
                    f"Classification: {classification}",
                    "Direction",
                    f"{node_name}<> {descs}",
                    f"Peers : {peers_str}" if peers_str else "Peers :",
                    f"Interface : {ifaces_str}",
                    # status placeholders will be in Excel; also include statuses from log (last_state) and int_status
                    "Status : " + " , ".join(f"{{{e['iface']} : {e['log_state']}}}" for e in entries),
                    "2nd line informed : No",
                    f"Alarm time: {alarm_time}",
                    ""
                ]
                txt_report_blocks.append("\n".join(block_lines))

            report_per_node[node_name] = node_entries
            conn.disconnect()
        except NetMikoTimeoutException as e:
            log(f"{node_name} - TIMEOUT: {e}")
            txt_report_blocks.append(f"{'-'*41}{node_name}{'-'*41}\nERROR: TIMEOUT connecting to node\n")
            report_per_node[node_name] = node_entries
        except NetMikoAuthenticationException as e:
            log(f"{node_name} - AUTH_FAIL: {e}")
            txt_report_blocks.append(f"{'-'*41}{node_name}{'-'*41}\nERROR: AUTH failure\n")
            report_per_node[node_name] = node_entries
        except Exception as e:
            log(f"{node_name} - ERROR: {e}")
            txt_report_blocks.append(f"{'-'*41}{node_name}{'-'*41}\nERROR: {e}\n")
            report_per_node[node_name] = node_entries

        # delay between nodes

    # write combined text file with header
    with open(TXT_FILE, "w", encoding="utf-8") as tf:
        tf.write("****************************************************************************\n")
        tf.write("*               ! ! ! ! ! !    A H M E D   F A R E S    ! ! ! ! ! !        * \n")
        tf.write("****************************************************************************\n\n")
        tf.write(f"Combined report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        for b in txt_report_blocks:
            tf.write(b + "\n\n")

    # write excel file
    write_excel(report_per_node)

    print(f"Done. Excel: {EXCEL_FILE}  Text: {TXT_FILE}  Log: {LOG_FILE}")

# ------------------------- Excel writer -------------------------
def write_excel(report_per_node):
    wb = Workbook()
    ws = wb.active
    ws.title = "BGP Summary"

    # Build header structure:
    # Columns: Node | [Etisalat: Interface, IP, Status, Description] | [Orange: ...] | [WE: ...] | [OTHER: ...]
    companies = ["Etisalat", "Orange", "WE", OTHER_COMPANY_NAME]
    subcols = ["Interface", "IP", "Status", "Description"]

    # Row 1: big header (title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(companies)*len(subcols))
    ws["A1"] = f"BGP Status Report- By : FARES - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    # Row 2: column group headers
    col = 1
    ws.cell(row=2, column=col, value="Node").font = BOLD
    ws.cell(row=2, column=col).alignment = CENTER
    col += 1
    for comp in companies:
        start = col
        end = col + len(subcols) - 1
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        ws.cell(row=2, column=start, value=comp).font = BOLD
        ws.cell(row=2, column=start).alignment = CENTER
        # set fill color per company
        fill = FILL_OTHER
        if comp == "Etisalat":
            fill = FILL_ETISALAT
        elif comp == "Orange":
            fill = FILL_ORANGE
        elif comp == "WE":
            fill = FILL_WE
        for c in range(start, end+1):
            ws.cell(row=2, column=c).fill = fill
        col = end + 1

    # Row 3: subcolumns
    col = 1
    ws.cell(row=3, column=col, value="").font = BOLD
    col += 1
    for comp in companies:
        for sc in subcols:
            ws.cell(row=3, column=col, value=sc).font = BOLD
            ws.cell(row=3, column=col).alignment = CENTER
            col += 1

    # Data rows: for each node, compute max rows needed (max number of interface entries across companies)
    row = 4
    for node_name, companies_map in report_per_node.items():
        # companies_map: company -> [entries]
        counts = [len(companies_map.get(c, [])) for c in companies]
        max_rows = max(counts) if counts else 1
        if max_rows == 0:
            max_rows = 1

        # Write node name in a merged cell spanning those rows in column A
        ws.merge_cells(start_row=row, start_column=1, end_row=row+max_rows-1, end_column=1)
        ws.cell(row=row, column=1, value=node_name).alignment = Alignment(vertical='top', horizontal='center')
        ws.cell(row=row, column=1).font = Font(bold=True)

        # fill each row with company entries if available
        for r_off in range(max_rows):
            col = 2
            for comp in companies:
                entries = companies_map.get(comp, [])
                if r_off < len(entries):
                    ent = entries[r_off]
                    iface = ent["iface"]
                    peer = ", ".join(ent["peers"]) if ent["peers"] else ""
                    # status: we will use log_state (UP/DOWN) as the main status (per your final decision)
                    status = ent.get("log_state", "UNKNOWN")
                    desc = ent.get("desc", "")
                    # write cells
                    ws.cell(row=row + r_off, column=col, value=iface)
                    ws.cell(row=row + r_off, column=col+1, value=peer)
                    ws.cell(row=row + r_off, column=col+2, value=status)
                    ws.cell(row=row + r_off, column=col+3, value=desc)
                    # style: border and alignment
                    for cc in range(col, col+4):
                        ws.cell(row=row + r_off, column=cc).border = THIN_BORDER
                        ws.cell(row=row + r_off, column=cc).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    # color status cell: red if DOWN, green if UP, else none
                    status_cell = ws.cell(row=row + r_off, column=col+2)
                    if str(status).upper() == "DOWN":
                        status_cell.fill = FILL_DOWN
                        # highlight entire entry row cells for visibility
                        for cc in range(col, col+4):
                            ws.cell(row=row + r_off, column=cc).fill = FILL_DOWN
                    elif str(status).upper() == "UP":
                        status_cell.fill = FILL_UP
                else:
                    # empty cells
                    for cc in range(col, col+4):
                        ws.cell(row=row + r_off, column=cc, value="")
                col += 4

        row += max_rows

    # Adjust column widths
    for i, width in enumerate([20] + [18]* (len(companies)*len(subcols)), start=1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = width

    # freeze panes
    ws.freeze_panes = ws['B4']

    # final save
    wb.save(EXCEL_FILE)

# ------------------------- Entrypoint -------------------------
if __name__ == "__main__":
    process_all_nodes()
